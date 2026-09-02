"""One-off ETL: flatten the JSA "Futures History.xlsx" workbook (per-contract daily
settlements, corn & soybeans, ~2006-2021) into a tidy CSV the app can ship and load fast.

Run again only if the source workbook is updated:
    python scripts/build_history_archive.py
"""

import re
import sys
from pathlib import Path

import openpyxl
import pandas as pd

SOURCE = (
    r"C:\Users\KoltenPostin\John Stewart and Associates\JSA - Documents"
    r"\Research Analyst\Misc\Future Seasonal Charts\Futures History.xlsx"
)
OUT = Path(__file__).parent.parent / "data" / "futures_history_archive.csv"

TICKER_RE = re.compile(r"^(ZC|ZS)([FGHJKMNQUVXZ])(\d{2})$")


def main():
    wb = openpyxl.load_workbook(SOURCE, data_only=True, read_only=True)
    rows = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        it = ws.iter_rows(values_only=True)
        header = next(it)
        columns = header[1:]
        parsed = [TICKER_RE.match(c or "") for c in columns]
        skipped = [c for c, m in zip(columns, parsed) if not m]
        if skipped:
            print(f"  {sheet_name}: skipping unrecognized columns {skipped}", file=sys.stderr)
        for row in it:
            trade_date = row[0]
            if trade_date is None:
                continue
            for col, m, price in zip(columns, parsed, row[1:]):
                if not m or price is None:
                    continue
                product, month, yy = m.groups()
                rows.append((product, month, 2000 + int(yy), trade_date.date(), float(price)))
        print(f"  {sheet_name}: {ws.max_row - 1} rows scanned")

    df = pd.DataFrame(rows, columns=["product_code", "month", "year", "date", "price"])
    df = df.sort_values(["product_code", "month", "year", "date"]).drop_duplicates(
        ["product_code", "month", "year", "date"]
    )
    OUT.parent.mkdir(exist_ok=True)
    df.to_csv(OUT, index=False)
    print(f"Wrote {len(df):,} rows -> {OUT}")
    print(df.groupby(["product_code", "month"])["year"].agg(["min", "max", "nunique"]))


if __name__ == "__main__":
    main()
